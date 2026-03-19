import csv
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.patient import Patient
from app.schemas.importing import ImportErrorItem
from app.schemas.patient import (
    BaselineFeatureBase,
    FollowupOutcomeBase,
    LightInterventionBase,
    PatientCreate,
    PatientUpdate,
    PredictionResultBase,
    QuestionnaireScoreBase,
    SleepMetricBase,
)
from app.services.audit_service import add_audit_log
from app.services.patient_service import create_patient, update_patient

settings = get_settings()

TEMPLATE_HEADERS = [
    "患者编号",
    "匿名编号",
    "性别",
    "年龄",
    "身高(cm)",
    "体重(kg)",
    "教育程度",
    "备注",
    "作息",
    "病程",
    "用药",
    "合并症",
    "心理状态",
    "睡眠习惯",
    "基线备注",
    "PSQI",
    "ISI",
    "焦虑评分",
    "抑郁评分",
    "评估日期",
    "总睡眠时间(h)",
    "入睡潜伏期(min)",
    "睡眠效率(%)",
    "觉醒次数",
    "睡眠指标备注",
    "光照强度(lux)",
    "开始时段",
    "持续时间(min)",
    "干预天数",
    "依从性",
    "不良反应",
    "随访日期",
    "主要结局",
    "次要结局",
    "随访备注",
    "推荐等级",
    "推荐评分",
    "解释文本",
    "模型版本",
    "模型类型",
    "模型状态",
    "模型说明",
    "生成时间",
]

EXAMPLE_ROW = {
    "患者编号": "P-0001",
    "匿名编号": "A-0001",
    "性别": "女",
    "年龄": "34",
    "身高(cm)": "162",
    "体重(kg)": "54",
    "教育程度": "硕士",
    "备注": "演示样例",
    "作息": "23:30-07:00",
    "病程": "18个月",
    "用药": "偶尔使用助眠药物",
    "合并症": "无",
    "心理状态": "轻度焦虑",
    "睡眠习惯": "睡前使用手机",
    "基线备注": "基线资料完整",
    "PSQI": "13",
    "ISI": "17",
    "焦虑评分": "9",
    "抑郁评分": "6",
    "评估日期": "2026-03-19",
    "总睡眠时间(h)": "5.4",
    "入睡潜伏期(min)": "48",
    "睡眠效率(%)": "72",
    "觉醒次数": "3",
    "睡眠指标备注": "近一周平均值",
    "光照强度(lux)": "2500",
    "开始时段": "07:30",
    "持续时间(min)": "30",
    "干预天数": "14",
    "依从性": "高",
    "不良反应": "无明显不良反应",
    "随访日期": "2026-04-02",
    "主要结局": "ISI下降4分",
    "次要结局": "睡眠效率提升",
    "随访备注": "依从性良好",
    "推荐等级": "高",
    "推荐评分": "0.86",
    "解释文本": "晨间中等强度光照，预期获益较高",
    "模型版本": "规则基线版",
    "模型类型": "rule",
    "模型状态": "active",
    "模型说明": "科研原型默认规则模型",
    "生成时间": "2026-03-19 09:00:00",
}

REQUIRED_COLUMNS = ["患者编号", "匿名编号"]
ALLOWED_MODEL_TYPES = {"rule", "predictive", "causal"}


@dataclass
class ImportExecutionResult:
    total_rows: int
    success_count: int
    failed_count: int
    created_count: int
    updated_count: int
    errors: list[ImportErrorItem]


def generate_template_file(file_format: str) -> tuple[bytes, str, str]:
    normalized = file_format.lower()
    if normalized == "csv":
        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=TEMPLATE_HEADERS)
        writer.writeheader()
        writer.writerow(EXAMPLE_ROW)
        return (
            buffer.getvalue().encode("utf-8-sig"),
            "text/csv; charset=utf-8",
            "patients_import_template.csv",
        )

    if normalized == "xlsx":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "患者导入模板"
        worksheet.append(TEMPLATE_HEADERS)
        worksheet.append([EXAMPLE_ROW[header] for header in TEMPLATE_HEADERS])
        output = BytesIO()
        workbook.save(output)
        return (
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "patients_import_template.xlsx",
        )

    raise ValueError("模板格式仅支持 csv 或 xlsx")


def _normalize_header(value) -> str:
    return str(value).strip() if value is not None else ""


def _parse_csv(file_bytes: bytes) -> list[dict[str, str]]:
    last_error = None
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = file_bytes.decode(encoding)
            reader = csv.DictReader(StringIO(text))
            return [{_normalize_header(k): (v or "").strip() for k, v in row.items()} for row in reader]
        except UnicodeDecodeError as exc:
            last_error = exc

    raise ValueError("CSV 文件编码无法识别，请使用 UTF-8 或 GBK 编码后重试") from last_error


def _parse_xlsx(file_bytes: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(cell) for cell in rows[0]]
    results = []
    for raw_row in rows[1:]:
        if all(value in (None, "") for value in raw_row):
            continue
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = raw_row[index] if index < len(raw_row) else ""
            row[header] = "" if value is None else str(value).strip()
        results.append(row)
    return results


def load_rows_from_upload(file_name: str, file_bytes: bytes) -> list[dict[str, str]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        rows = _parse_csv(file_bytes)
    elif suffix == ".xlsx":
        rows = _parse_xlsx(file_bytes)
    else:
        raise ValueError("仅支持导入 CSV 或 XLSX 文件")

    if not rows:
        raise ValueError("导入文件为空，请先填写模板内容后再上传")

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in rows[0]]
    if missing_columns:
        raise ValueError(f"导入模板缺少必填列：{'、'.join(missing_columns)}")

    return rows


def _parse_int(value: str | None, field_name: str, errors: list[str]) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except ValueError:
        errors.append(f"{field_name}不是有效整数")
        return None


def _parse_float(value: str | None, field_name: str, errors: list[str]) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except ValueError:
        errors.append(f"{field_name}不是有效数字")
        return None


def _parse_date(value: str | None, field_name: str, errors: list[str]) -> date | None:
    if value in (None, ""):
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    errors.append(f"{field_name}日期格式应为 YYYY-MM-DD")
    return None


def _parse_datetime(value: str | None, field_name: str, errors: list[str]) -> datetime | None:
    if value in (None, ""):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    errors.append(f"{field_name}时间格式应为 YYYY-MM-DD HH:MM:SS")
    return None


def _build_patient_payload(row: dict[str, str]) -> tuple[PatientCreate, list[str]]:
    errors: list[str] = []

    patient_code = row.get("患者编号", "").strip()
    anonymized_code = row.get("匿名编号", "").strip()
    if not patient_code:
        errors.append("患者编号不能为空")
    if not anonymized_code:
        errors.append("匿名编号不能为空")

    model_type = row.get("模型类型", "").strip()
    if model_type and model_type not in ALLOWED_MODEL_TYPES:
        errors.append("模型类型仅支持 rule、predictive、causal")

    payload = PatientCreate(
        patient_code=patient_code or "TEMP",
        anonymized_code=anonymized_code or "TEMP",
        gender=row.get("性别") or None,
        age=_parse_int(row.get("年龄"), "年龄", errors),
        height_cm=_parse_float(row.get("身高(cm)"), "身高(cm)", errors),
        weight_kg=_parse_float(row.get("体重(kg)"), "体重(kg)", errors),
        education_level=row.get("教育程度") or None,
        remarks=row.get("备注") or None,
        baseline_feature=BaselineFeatureBase(
            work_rest_schedule=row.get("作息") or None,
            disease_duration=row.get("病程") or None,
            medication_usage=row.get("用药") or None,
            comorbidities=row.get("合并症") or None,
            psychological_status=row.get("心理状态") or None,
            sleep_habits=row.get("睡眠习惯") or None,
            notes=row.get("基线备注") or None,
        ),
        questionnaire_score=QuestionnaireScoreBase(
            psqi_score=_parse_float(row.get("PSQI"), "PSQI", errors),
            isi_score=_parse_float(row.get("ISI"), "ISI", errors),
            anxiety_score=_parse_float(row.get("焦虑评分"), "焦虑评分", errors),
            depression_score=_parse_float(row.get("抑郁评分"), "抑郁评分", errors),
            assessed_at=_parse_date(row.get("评估日期"), "评估日期", errors),
        ),
        sleep_metric=SleepMetricBase(
            total_sleep_time_hours=_parse_float(row.get("总睡眠时间(h)"), "总睡眠时间(h)", errors),
            sleep_latency_minutes=_parse_float(row.get("入睡潜伏期(min)"), "入睡潜伏期(min)", errors),
            sleep_efficiency=_parse_float(row.get("睡眠效率(%)"), "睡眠效率(%)", errors),
            awakening_count=_parse_int(row.get("觉醒次数"), "觉醒次数", errors),
            notes=row.get("睡眠指标备注") or None,
        ),
        light_intervention=LightInterventionBase(
            intensity_lux=_parse_float(row.get("光照强度(lux)"), "光照强度(lux)", errors),
            start_period=row.get("开始时段") or None,
            duration_minutes=_parse_int(row.get("持续时间(min)"), "持续时间(min)", errors),
            intervention_days=_parse_int(row.get("干预天数"), "干预天数", errors),
            adherence=row.get("依从性") or None,
            adverse_events=row.get("不良反应") or None,
        ),
        followup_outcome=FollowupOutcomeBase(
            followup_date=_parse_date(row.get("随访日期"), "随访日期", errors),
            primary_outcome=row.get("主要结局") or None,
            secondary_outcome=row.get("次要结局") or None,
            notes=row.get("随访备注") or None,
        ),
        prediction_result=PredictionResultBase(
            recommendation_level=row.get("推荐等级") or None,
            score=_parse_float(row.get("推荐评分"), "推荐评分", errors),
            explanation_text=row.get("解释文本") or None,
            model_version_name=row.get("模型版本") or None,
            model_version_type=model_type or None,
            model_version_status=row.get("模型状态") or None,
            model_version_description=row.get("模型说明") or None,
            generated_at=_parse_datetime(row.get("生成时间"), "生成时间", errors),
        ),
    )

    return payload, errors


def import_patients(
    db: Session,
    *,
    file_name: str,
    rows: list[dict[str, str]],
    actor_name: str | None = None,
) -> ImportExecutionResult:
    errors: list[ImportErrorItem] = []
    created_count = 0
    updated_count = 0
    success_count = 0
    actor = actor_name or settings.demo_username

    for index, row in enumerate(rows, start=2):
        patient_code = (row.get("患者编号") or "").strip() or None
        payload, row_errors = _build_patient_payload(row)
        if row_errors:
            errors.append(
                ImportErrorItem(
                    row_number=index,
                    patient_code=patient_code,
                    message="；".join(row_errors),
                )
            )
            continue

        try:
            existing = db.scalar(select(Patient).where(Patient.patient_code == payload.patient_code))
            if existing is None:
                create_patient(db, payload)
                created_count += 1
            else:
                update_patient(
                    db,
                    existing,
                    PatientUpdate(
                        anonymized_code=payload.anonymized_code,
                        gender=payload.gender,
                        age=payload.age,
                        height_cm=payload.height_cm,
                        weight_kg=payload.weight_kg,
                        education_level=payload.education_level,
                        remarks=payload.remarks,
                        baseline_feature=payload.baseline_feature,
                        questionnaire_score=payload.questionnaire_score,
                        sleep_metric=payload.sleep_metric,
                        light_intervention=payload.light_intervention,
                        followup_outcome=payload.followup_outcome,
                        prediction_result=payload.prediction_result,
                    ),
                )
                updated_count += 1
            success_count += 1
        except Exception as exc:
            db.rollback()
            errors.append(
                ImportErrorItem(
                    row_number=index,
                    patient_code=patient_code,
                    message=f"导入失败：{exc}",
                )
            )

    result = ImportExecutionResult(
        total_rows=len(rows),
        success_count=success_count,
        failed_count=len(errors),
        created_count=created_count,
        updated_count=updated_count,
        errors=errors,
    )

    add_audit_log(
        db,
        actor_name=actor,
        action_type="import_patients",
        target_type="patients",
        details={
            "file_name": file_name,
            "total_rows": result.total_rows,
            "success_count": result.success_count,
            "failed_count": result.failed_count,
            "created_count": result.created_count,
            "updated_count": result.updated_count,
            "error_preview": [error.message for error in result.errors[:5]],
        },
        detail_text=f"导入患者文件 {file_name}",
    )
    db.commit()
    return result
